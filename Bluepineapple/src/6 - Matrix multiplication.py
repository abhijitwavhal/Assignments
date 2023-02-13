# Program to find resultant matrix after multiplying two matrices


import openpyxl
from tabulate import tabulate

def matrixInput(row, column, elements):
    matrix = []
    rowElements = []

    elements = elements.split(",")

    for iterator in elements:
        rowElements.append(int(iterator))

    for iterator in range(0, len(rowElements), column):
        matrix.append(rowElements[iterator: iterator + column])

    return matrix

def displayMatrix(row, column, matrix):
    for iterator1 in range(row):
        for iterator2 in range(column):
            print(matrix[iterator1][iterator2], end = " ")
        print()

def matrixMultiplication(rowA, matrixA, columnB, matrixB, matrixC):
    if rowA == columnB:
        for iterator1 in range(len(matrixA)):
            for iterator3 in range(len(matrixB[0])):
                for iterator2 in range(len(matrixB)):
                    matrixC[iterator1][iterator3] = matrixC[iterator1][iterator3] + (matrixA[iterator1][iterator2] * matrixB[iterator2][iterator3])
        return matrixC

    else:
        print("\nMatrix multiplication is not possible")

def testMatrixMultiplication():
        workbook = openpyxl.load_workbook("test.xlsx")

        sheet = workbook.active

        table = [['Sr.No.', 'Input (Matrix A)', 'Input (Matrix B)', 'Expected Output', 'Actual Output', 'Status']]

        for iterator1 in range(2, sheet.max_row + 1):
            for iterator2 in range(1, sheet.max_column + 1):
                rowA = int((sheet.cell(row=iterator1, column=1)).value)
                columnA = int((sheet.cell(row=iterator1, column=2)).value)
                elementsA = str((sheet.cell(row=iterator1, column=3)).value)

                rowB = int((sheet.cell(row=iterator1, column=4)).value)
                columnB = int((sheet.cell(row=iterator1, column=5)).value)
                elementsB = str((sheet.cell(row=iterator1, column=6)).value)

                rowC = int((sheet.cell(row=iterator1, column=7)).value)
                columnC = int((sheet.cell(row=iterator1, column=8)).value)
                elementsC = str((sheet.cell(row=iterator1, column=9)).value)

                expectedOutput = str((sheet.cell(row=iterator1, column=10)).value)

                matrixA = matrixInput(rowA, columnA, elementsA)
                matrixB = matrixInput(rowB, columnB, elementsB)
                matrixC = matrixInput(rowC, columnC, elementsC)

                expectedOutput = matrixInput(rowC, columnC, expectedOutput)

            matrixC = matrixMultiplication(rowA, matrixA, columnB, matrixB, matrixC)

            if (expectedOutput == matrixC):
                table.append([iterator1-1, matrixA, matrixB, expectedOutput, matrixC, "passed"])

            else:
                table.append([iterator1 - 1, matrixA, matrixB, expectedOutput, matrixC, "failed"])

        print(tabulate(table, headers='firstrow', tablefmt='fancy_grid'))

def main():
    testMatrixMultiplication()

if __name__ == "__main__":
    main()